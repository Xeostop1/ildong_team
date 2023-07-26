import openpyxl


def find_cells_with_two_words(input_file, output_file, word1, word2):
    # 입력 파일을 엽니다.
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active

    # 출력 파일에서 사용할 워크북과 워크시트를 생성합니다.
    output_wb = openpyxl.Workbook()
    output_ws = output_wb.active

    # 엑셀 워크시트를 순회하며 두 단어가 있는 셀을 찾습니다.
    row_output = 1
    for row in ws.iter_rows():
        for cell in row:
            if word1 in str(cell.value) and word2 in str(cell.value):
                output_ws.cell(row=row_output, column=1, value=cell.coordinate)
                output_ws.cell(row=row_output, column=2, value=cell.value)
                row_output += 1

    # 두 단어를 포함하는 셀 정보를 새 엑셀 파일로 저장합니다.
    output_wb.save(output_file)


# 사용 예
word1 = "협찬"
word2 = "기미"
input_file = "./pre_foler/gimi_origin.xlsx"  # 원본 엑셀 파일 경로
# output_file = f"{word1}_{word2}.xlsx"  # 출력 엑셀 파일 경로
output_file = f"{word1}_{word2}.txt"  # 출력 엑셀 파일 경로


find_cells_with_two_words(input_file, output_file, word1, word2)
print("완료")
